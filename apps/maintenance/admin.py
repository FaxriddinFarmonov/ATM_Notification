import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib import admin
from django.http import HttpResponse
from .models import DocumentUpload, MaintenanceProtocol, MaintenanceItem


@admin.action(description="Tanlangan qatorlarni Excel formatida yuklab olish")
def export_to_excel_action(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="maintenance_report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "T/r", "Modul", "Seriya raqami", "Ehtiyot qismi / Xizmat",
        "Miqdori", "Narxi", "Jami", "QQS", "Jami (QQS bilan)", "Filial", "MFO"
    ]
    ws.append(headers)
    ws.row_dimensions[1].height = 26

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    if queryset.model == MaintenanceItem:
        items = queryset
    elif queryset.model == DocumentUpload:
        items = MaintenanceItem.objects.filter(document__in=queryset)
    else:
        items = MaintenanceItem.objects.filter(protocol__in=queryset)

    center_cols = {1, 11}        
    left_cols = {2, 3, 4, 10}
    number_cols = {5, 6, 7, 8, 9}

    for item in items:
        row_data = [
            item.row_number, item.equipment_module, item.serial_number, item.part_name,
            float(item.quantity), float(item.price_per_unit),
            float(item.total_amount), float(item.vat_amount), float(item.total_with_vat),
            item.filial_name, item.mfo_bank,
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=curr_row, column=col_idx)
            if col_idx in center_cols:
                c.alignment = center_align
            elif col_idx in left_cols:
                c.alignment = left_align
            elif col_idx in number_cols:
                c.alignment = right_align
                c.number_format = '#,##0.00'
            c.font = data_font
            c.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(response)
    return response


class MaintenanceItemInline(admin.TabularInline):
    model = MaintenanceItem
    extra = 0 
    max_num = 0 
    can_delete = False

    fields = [
        'row_number', 'equipment_module', 'serial_number', 'part_name',
        'quantity', 'price_per_unit', 'total_amount', 'vat_amount', 'total_with_vat',
        'filial_name', 'mfo_bank',
    ]
    readonly_fields = fields

    class Media:
        css = {
            'all': ('admin/css/forms.css',)
        }


@admin.register(DocumentUpload)
class DocumentUploadAdmin(admin.ModelAdmin):
    list_display = ['title', 'get_protocol_date', 'is_processed']
    list_filter = ['is_processed', 'uploaded_at']
    
    search_fields = [
        'title', 
        'maintenanceitem__filial_name', 
        'maintenanceitem__part_name', 
        'maintenanceitem__serial_number'
    ]
    inlines = [MaintenanceItemInline]
    actions = [export_to_excel_action]

    @admin.display(description="Hujjat sanasi")
    def get_protocol_date(self, obj):
        protocol = MaintenanceProtocol.objects.filter(document_source=obj).first()
        if protocol and protocol.protocol_date:
            return protocol.protocol_date.strftime('%d.%m.%Y')
        return "Noma'lum"


@admin.register(MaintenanceProtocol)
class MaintenanceProtocolAdmin(admin.ModelAdmin):
    list_display = ['protocol_number', 'protocol_date', 'performer_company', 'customer_bank']
    list_filter = ['protocol_date', 'performer_company']
    search_fields = ['protocol_number', 'performer_company', 'customer_bank']
    actions = [export_to_excel_action]


@admin.register(MaintenanceItem)
class MaintenanceItemAdmin(admin.ModelAdmin):
    list_display = [
        'row_number', 'get_document_with_date', 'equipment_module', 'serial_number', 'part_name',
        'quantity', 'price_per_unit', 'total_with_vat', 'filial_name', 'mfo_bank',
    ]
    list_display_links = ['get_document_with_date']
    list_select_related = ['document', 'protocol']
    list_per_page = 30

    list_filter = [
        'document__title', 
        'protocol_date',
        'equipment_module', 
        'filial_name'
    ]
    search_fields = [
        'part_name', 'serial_number', 'filial_name', 
        'protocol__protocol_number', 'document__title'
    ]
    actions = [export_to_excel_action]

    @admin.display(description="Hujjat")
    def get_document_with_date(self, obj):
        doc_title = obj.document.title if obj.document and obj.document.title else "Hujjat"
        if obj.protocol_date:
            return f"{doc_title} ({obj.protocol_date.strftime('%d.%m.%Y')})"
        return doc_title