from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db.models import Sum, Q
from apps.maintenance.models import MaintenanceItem
from collections import defaultdict

from openpyxl.drawing.image import Image
from pathlib import Path

from ..models import ATMServicePayment
from ..models import (
    ATMMonthlyStatistic,
    ATMYearStatistic,
)

class ATMExcelExporter:
    """
    Professional ATM Excel Exporter.

    Creates a complete ATM report including:

        • ATM information
        • Monthly statistics
        • Year statistics

    Output:
        BytesIO
    """

    SHEET_NAME = "ATM Report"

    def __init__(self, atm):

        self.atm = atm

        self.workbook = Workbook()

        self.sheet = self.workbook.active

        self.sheet.title = self.SHEET_NAME

    def export(self):
        """
        Excel faylni yaratadi va BytesIO obyektini qaytaradi.
        """

        # Workbook properties
        self.prepare_document()
        self.build_service_cache()
        # 3 ta sheet yaratamiz
        self.create_sheets()

        # ATM ma'lumotlari
        self.write_information()

        # Oylik statistika
        self.write_month_statistics()

        # Yillik statistika
        self.write_year_statistics()

        # Har bir sheet uchun ustunlarni moslashtiramiz
        for ws in self.workbook.worksheets:

            for column in ws.columns:

                length = 0

                letter = get_column_letter(column[0].column)

                for cell in column:

                    value = "" if cell.value is None else str(cell.value)

                    if len(value) > length:
                        length = len(value)

                ws.column_dimensions[letter].width = min(length + 3, 40)

        # Freeze
        self.info_sheet.freeze_panes = "A1"
        self.month_sheet.freeze_panes = "A2"
        self.year_sheet.freeze_panes = "A2"

        # BytesIO qaytaramiz
        return self.save()

    def prepare_document(self):

        props = self.workbook.properties

        props.creator = "Turonbank"

        props.lastModifiedBy = "Turonbank"

        props.company = "Turonbank"

        props.title = "ATM Report"

        props.subject = "ATM Statistics"

        props.description = (
            "ATM Monthly and Year Statistics"
        )

        props.created = datetime.now()




    def create_sheets(self):

        self.info_sheet = self.workbook.active

        self.info_sheet.title = "ATM Information"

        self.month_sheet = self.workbook.create_sheet(
            title="Monthly Statistics"
        )

        self.year_sheet = self.workbook.create_sheet(
            title="Year Statistics"
        )

    def build_service_cache(self):
        """
        Build service cache.

        Business rules

        1. Incassation / Rent / Electricity
           - Faqat bazadagi (year, month) uchun yoziladi.
           - Hech qanday keyingi oylarga tarqatilmaydi.

        2. BTECH / GLOB
           - Contract qiymati.
           - Eng birinchi payment sanasidan boshlab
             statistikada mavjud barcha oylarga yoziladi.

        3. Year cache
           - Month cache yig'indisidan avtomatik hisoblanadi.
        """

        from collections import defaultdict

        self.month_service_cache = {}
        self.year_service_cache = {}

        contract = getattr(
            self.atm,
            "service_contract",
            None,
        )

        if not contract:
            return

        payments = list(
            contract.payments.all().order_by(
                "year",
                "month",
                "payment_type",
            )
        )

        if not payments:
            return

        monthly = {}
        yearly = {}

        # -------------------------------------------------------
        # Contract boshlanish sanasi
        # -------------------------------------------------------

        first_payment = min(
            payments,
            key=lambda p: (
                p.year,
                p.month,
            ),
        )

        contract_start = (
            first_payment.year,
            first_payment.month,
        )

        # -------------------------------------------------------
        # Statistikada mavjud oylarni yaratamiz
        # -------------------------------------------------------

        statistics = (
            ATMMonthlyStatistic.objects
            .filter(
                atm=self.atm,
            )
            .order_by(
                "year",
                "month",
            )
        )

        for stat in statistics:

            key = (
                stat.year,
                stat.month,
            )

            monthly[key] = {
                "btech": 0,
                "glob": 0,
                "incassation": 0,
                "rent": 0,
                "electricity": 0,
            }

            # Contract boshlangan bo'lsa
            if key >= contract_start:
                monthly[key]["btech"] = float(
                    contract.btech_monthly_fee or 0
                )

                monthly[key]["glob"] = float(
                    contract.glob_monthly_fee or 0
                )

        # -------------------------------------------------------
        # Paymentlarni faqat o'z oyiga yozamiz
        # -------------------------------------------------------

        for payment in payments:

            key = (
                payment.year,
                payment.month,
            )

            if key not in monthly:
                continue

            if payment.payment_type == ATMServicePayment.PaymentType.INCASSATION:

                monthly[key]["incassation"] = float(
                    payment.amount
                )

            elif payment.payment_type == ATMServicePayment.PaymentType.RENT:

                monthly[key]["rent"] = float(
                    payment.amount
                )

            elif payment.payment_type == ATMServicePayment.PaymentType.ELECTRICITY:

                monthly[key]["electricity"] = float(
                    payment.amount
                )

        # -------------------------------------------------------
        # Year cache
        # -------------------------------------------------------

        for (year, month), values in monthly.items():

            if year not in yearly:
                yearly[year] = {
                    "btech": 0,
                    "glob": 0,
                    "incassation": 0,
                    "rent": 0,
                    "electricity": 0,
                }

            yearly[year]["btech"] += values["btech"]
            yearly[year]["glob"] += values["glob"]
            yearly[year]["incassation"] += values["incassation"]
            yearly[year]["rent"] += values["rent"]
            yearly[year]["electricity"] += values["electricity"]

        self.month_service_cache = monthly

        self.year_service_cache = yearly
    def write_information(self):

        ws = self.info_sheet

        tech = getattr(self.atm, "technical", None)

        # =========================
        # LOGO
        # =========================

        logo_path = Path("templates/turonbank.png")

        if logo_path.exists():
            logo = Image(str(logo_path))

            logo.width = 220
            logo.height = 70

            ws.add_image(logo, "A1")

        # =========================
        # TITLE
        # =========================

        ws.merge_cells("C1:H2")

        title = ws["C1"]
        #
        # title.value = "TURONBANK ATM REPORT"

        title.font = Font(
            size=20,
            bold=True,
            color="FFFFFF",
        )

        title.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # title.fill = PatternFill(
        #     fill_type="solid",
        #     start_color="1F4E78",
        #     end_color="1F4E78",
        # )

        # =========================
        # EXPORT INFO
        # =========================

        ws["A5"] = "Generated"
        ws["B5"] = datetime.now().strftime("%d.%m.%Y %H:%M")

        ws["A6"] = "Generated By"
        ws["B6"] = "Turonbank Monitoring System"

        ws["A5"].font = Font(bold=True)
        ws["A6"].font = Font(bold=True)

        row = 8

        # =========================
        # GENERAL INFORMATION
        # =========================

        ws.merge_cells(f"A{row}:B{row}")

        header = ws[f"A{row}"]

        header.value = "GENERAL INFORMATION"

        header.font = Font(
            bold=True,
            color="FFFFFF",
            size=13,
        )

        header.fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
            end_color="4F81BD",
        )

        row += 2

        data = [

            ("Region", self.atm.region),

            ("ATM Name", self.atm.name),

            ("Address", self.atm.address),

            ("PROCESSING", self.atm.card_type),

            ("ATM Model", self.atm.model),

            (
                "Purchase Date",
                tech.purchase_date.strftime("%d.%m.%Y")
                if tech and tech.purchase_date
                else "",
            ),

            (
                "Purchase Price",
                float(tech.purchase_price)
                if tech and tech.purchase_price
                else "",
            ),

        ]
        thin = Side(style="thin")

        for key, value in data:
            ws.cell(row=row, column=1).value = key

            ws.cell(row=row, column=2).value = value

            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=1).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

            ws.cell(row=row, column=2).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

            row += 1

        row += 2

        # =========================
        # TECHNICAL INFORMATION
        # =========================

        ws.merge_cells(f"A{row}:B{row}")

        header = ws[f"A{row}"]

        header.value = "TECHNICAL INFORMATION"

        header.font = Font(
            bold=True,
            color="FFFFFF",
            size=13,
        )

        header.fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD",
            end_color="4F81BD",
        )

        row += 2

        technical = [

            ("Merchant ID", tech.merchant_id if tech else ""),
            ("Terminal ID", self.atm.terminal_id),
            ("Status", tech.status if tech else ""),

            ("Serial Number", tech.serial_number if tech else ""),

            ("Inventory Number", tech.inventory_number if tech else ""),

            ("23510 Account", tech.account_23510 if tech else ""),

            ("45265 Account", tech.account_45265 if tech else ""),

        ]

        for key, value in technical:
            ws.cell(row=row, column=1).value = key

            ws.cell(row=row, column=2).value = value

            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=1).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

            ws.cell(row=row, column=2).border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin,
            )

            row += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 55

    def write_month_statistics(self):

        ws = self.month_sheet

        ws.append([
            "Year",
            "Month",

            "Income",
            "Cash Withdraw",

            "Repair Cost",

            "Quantity",

            "BTECH Service Monthly Fee",
            "GLOB Service Monthly Fee",

            "Incassation Payment",
            "Rent Payment",
            "Electricity Payment",

            "Status",
            "Serial Number",
            "Merchant ID",
            "Inventory Number",
            "Purchase Date",
            "Purchase Price",
        ])
        totals = {
            "income": 0.0,
            "expense": 0.0,
            "repair": 0.0,
            "quantity": 0.0,
            "btech": 0.0,
            "glob": 0.0,
            "incassation": 0.0,
            "rent": 0.0,
            "electricity": 0.0,
        }

        statistics = (
            ATMMonthlyStatistic.objects
            .filter(
                atm=self.atm,
            )
            .order_by(
                "year",
                "month",
            )
        )

        tech = getattr(
            self.atm,
            "technical",
            None,
        )

        contract = getattr(
            self.atm,
            "service_contract",
            None,
        )

        contract_year = None

        if contract:

            first_payment = (
                contract.payments
                .order_by(
                    "year",
                    "month",
                )
                .first()
            )

            if first_payment:
                contract_year = first_payment.year

        months = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }

        for item in statistics:

            repair = 0
            quantity = 0

            if tech:
                repairs = (
                    MaintenanceItem.objects
                    .filter(
                        technical=tech,
                        protocol_date__year=item.year,
                        protocol_date__month=item.month,
                    )
                )

                repair = (
                        repairs.aggregate(
                            total=Sum(
                                "total_with_vat",
                            ),
                        )["total"]
                        or 0
                )

                quantity = (
                        repairs.aggregate(
                            total=Sum(
                                "quantity",
                            ),
                        )["total"]
                        or 0
                )

            service = self.month_service_cache.get(
                (
                    item.year,
                    item.month,
                ),
                {},
            )
            service = self.month_service_cache.get(
                (
                    item.year,
                    item.month,
                ),
                {},
            )

            btech = float(service.get("btech", 0))
            glob = float(service.get("glob", 0))

            incassation = float(service.get("incassation", 0))
            rent = float(service.get("rent", 0))
            electricity = float(service.get("electricity", 0))

            ws.append([
                item.year,

                months.get(item.month, item.month),

                float(item.income),

                float(item.expense),

                float(repair),

                float(quantity),

                btech,

                glob,

                incassation,

                rent,

                electricity,

                tech.status if tech else "",

                tech.serial_number if tech else "",

                tech.merchant_id if tech else "",

                tech.inventory_number if tech else "",
                tech.purchase_date.strftime("%d.%m.%Y")
                if tech and tech.purchase_date
                else "",

                float(tech.purchase_price)
                if tech and tech.purchase_price
                else "",
            ])
            totals["income"] += float(item.income)
            totals["expense"] += float(item.expense)
            totals["repair"] += float(repair)
            totals["quantity"] += float(quantity)
            totals["btech"] += btech
            totals["glob"] += glob
            totals["incassation"] += incassation
            totals["rent"] += rent
            totals["electricity"] += electricity
        ws.append([
            "TOTAL",
            "",

            totals["income"],
            totals["expense"],

            totals["repair"],

            totals["quantity"],

            totals["btech"],
            totals["glob"],

            totals["incassation"],
            totals["rent"],
            totals["electricity"],

            "",
            "",
            "",
            "",
        ])

    def write_year_statistics(self):

        ws = self.year_sheet

        ws.append([
            "Year",

            "Income",
            "Cash Withdraw",

            "Year Repair Cost",

            "Quantity",

            "BTECH Service Year Fee",
            "GLOB Service Year Fee",

            "Incassation Total",
            "Rent Total",
            "Electricity Total",

            "Status",
            "Serial Number",
            "Merchant ID",
            "Inventory Number",
            "Purchase Date",
            "Purchase Price",
        ])
        totals = {
            "income": 0.0,
            "expense": 0.0,
            "repair": 0.0,
            "quantity": 0.0,
            "btech": 0.0,
            "glob": 0.0,
            "incassation": 0.0,
            "rent": 0.0,
            "electricity": 0.0,
        }

        statistics = (
            ATMYearStatistic.objects
            .filter(
                atm=self.atm,
            )
            .order_by(
                "year",
            )
        )

        tech = getattr(
            self.atm,
            "technical",
            None,
        )

        for item in statistics:

            repair = 0
            quantity = 0

            if tech:
                repairs = (
                    MaintenanceItem.objects
                    .filter(
                        technical=tech,
                        protocol_date__year=item.year,
                    )
                )

                repair = (
                        repairs.aggregate(
                            total=Sum("total_with_vat"),
                        )["total"]
                        or 0
                )

                quantity = (
                        repairs.aggregate(
                            total=Sum("quantity"),
                        )["total"]
                        or 0
                )

            # -----------------------------------------
            # Service cache
            # -----------------------------------------

            service = self.year_service_cache.get(
                item.year,
                {},
            )

            btech = float(
                service.get(
                    "btech",
                    0,
                )
            )

            glob = float(
                service.get(
                    "glob",
                    0,
                )
            )

            incassation = float(
                service.get(
                    "incassation",
                    0,
                )
            )

            rent = float(
                service.get(
                    "rent",
                    0,
                )
            )

            electricity = float(
                service.get(
                    "electricity",
                    0,
                )
            )

            ws.append([
                item.year,

                float(item.income),
                float(item.expense),

                float(repair),
                float(quantity),

                btech,
                glob,

                incassation,
                rent,
                electricity,

                tech.status if tech else "",
                tech.serial_number if tech else "",
                tech.merchant_id if tech else "",
                tech.inventory_number if tech else "",

                tech.purchase_date.strftime("%d.%m.%Y")
                if tech and tech.purchase_date
                else "",

                float(tech.purchase_price)
                if tech and tech.purchase_price
                else "",
            ])
            totals["income"] += float(item.income)
            totals["expense"] += float(item.expense)
            totals["repair"] += float(repair)
            totals["quantity"] += float(quantity)
            totals["btech"] += btech
            totals["glob"] += glob
            totals["incassation"] += incassation
            totals["rent"] += rent
            totals["electricity"] += electricity
        ws.append([])

        ws.append([
            "TOTAL",

            totals["income"],
            totals["expense"],

            totals["repair"],
            totals["quantity"],

            totals["btech"],
            totals["glob"],

            totals["incassation"],
            totals["rent"],
            totals["electricity"],

            "",
            "",
            "",
            "",
        ])
    def auto_fit_columns(self):

        for column in self.sheet.columns:

            length = 0

            letter = get_column_letter(column[0].column)

            for cell in column:

                try:

                    value = str(cell.value)

                except Exception:

                    value = ""

                if len(value) > length:
                    length = len(value)

            self.sheet.column_dimensions[
                letter
            ].width = min(length + 3, 40)

    def build_response(self):

        output = self.export()

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response[
            "Content-Disposition"
        ] = (
            f'attachment; filename="ATM_{self.atm.terminal_id}.xlsx"'
        )

        return response

    def save(self):

        output = BytesIO()

        self.workbook.save(output)

        output.seek(0)

        return output

    def freeze(self):

        self.sheet.freeze_panes = "A9"