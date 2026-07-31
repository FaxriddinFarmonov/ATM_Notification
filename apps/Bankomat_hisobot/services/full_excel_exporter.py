import datetime
from pathlib import Path
from io import BytesIO
from collections import defaultdict
from openpyxl.styles import Font, PatternFill

from apps.maintenance.models import MaintenanceItem
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from django.db.models import Prefetch

from ..models import (
    ATMTURON,
    ATMTechnical,
    ATMMonthlyStatistic,
    ATMYearStatistic,
    ATMServicePayment
)


class FullATMExcelExporter:
    """
    Export every ATM into one professional Excel report.

    Workbook:

        1. Summary
        2. ATM Information
        3. Monthly Statistics
        4. Year Statistics
    """

    def __init__(self, queryset=None):

        # 1. Technical cache
        self.technical_cache = {
            tech.terminal_id.strip().upper(): tech
            for tech in ATMTechnical.objects.all()
        }

        # 2. Maintenance cache
        self.maintenance_cache = defaultdict(list)

        items = (
            MaintenanceItem.objects
            .select_related("technical")
            .exclude(technical=None)
        )

        for item in items:
            key = (
                item.technical_id,
                item.protocol_date.year,
                item.protocol_date.month,
            )
            self.maintenance_cache[key].append(item)

        # 3. Workbook
        self.workbook = Workbook()

        self.summary_sheet = self.workbook.active
        self.summary_sheet.title = "Summary"

        self.info_sheet = self.workbook.create_sheet("ATM Information")
        self.month_sheet = self.workbook.create_sheet("Monthly Statistics")
        self.year_sheet = self.workbook.create_sheet("Year Statistics")

        # 4. ATM lar
        if queryset is None:

            self.atms = (
                ATMTURON.objects
                .select_related("technical")
                .prefetch_related(
                    Prefetch(
                        "service_contract__payments",
                    ),
                    Prefetch(
                        "monthly_statistics",
                        queryset=ATMMonthlyStatistic.objects.order_by(
                            "year",
                            "month",
                        ),
                    ),
                    Prefetch(
                        "year_statistics",
                        queryset=ATMYearStatistic.objects.order_by(
                            "year",
                        ),
                    ),
                )
                .order_by(
                    "region",
                    "terminal_id",
                )
            )

        else:

            self.atms = queryset

        # 5. Service cache
        self.service_month_cache = {}
        self.service_year_cache = {}

        self.build_service_cache()
    def export(self):

        self.prepare_document()

        self.add_logo()

        self.write_summary()

        self.write_information()

        self.write_month_statistics()

        self.write_year_statistics()

        self.style_workbook()

        self.auto_fit()

        self.freeze()

        return self.save()

    def add_logo(self):

        logo_path = Path("templates/turonbank.png")

        if not logo_path.exists():
            return

        logo = Image(str(logo_path))

        # Logo o'lchami (balansli)
        logo.width = 300
        logo.height = 45

        # Logo joylashuvi
        self.summary_sheet.add_image(
            logo,
            "A1",
        )

        # 1-qator balandligi
        self.summary_sheet.row_dimensions[1].height = 42
    def style_workbook(self):
        """
        Workbook'ni professional ko'rinishga keltiradi.
        """

        # Ranglar
        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        summary_fill = PatternFill(
            fill_type="solid",
            start_color="DCEEFF",
            end_color="DCEEFF",
        )

        white_fill = PatternFill(fill_type=None)

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        bold_font = Font(
            bold=True,
        )

        thin = Side(
            style="thin",
            color="D9D9D9",
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        for sheet in self.workbook.worksheets:

            # Border + Alignment
            for row in sheet.iter_rows():

                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical="center"
                    )

            # ===========================
            # SUMMARY SHEET
            # ===========================

            if sheet.title == "Summary":

                # 1-qator mutlaqo oq bo'lsin
                for cell in sheet[1]:
                    cell.fill = white_fill
                    cell.font = bold_font
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )

                # Dashboard chap ustuni
                for row in range(8, sheet.max_row + 1):
                    sheet.cell(row, 1).fill = summary_fill
                    sheet.cell(row, 1).font = bold_font

                    sheet.cell(row, 2).font = Font(
                        bold=True,
                    )

                continue

            # ===========================
            # QOLGAN SHEETLAR
            # ===========================

            for cell in sheet[1]:
                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            # Header balandligi
            sheet.row_dimensions[1].height = 24

            for sheet in (self.month_sheet, self.year_sheet):

                row = sheet.max_row

                for cell in sheet[row]:
                    cell.font = Font(
                        bold=True,
                    )

                    cell.fill = PatternFill(
                        fill_type="solid",
                        start_color="FFF2CC",
                        end_color="FFF2CC",
                    )
    def prepare_document(self):
        props = self.workbook.properties

        props.creator = "Turonbank"

        props.company = "Turonbank"

        props.title = "ATM Full Report"

        props.subject = "ATM Statistics"

        props.description = (
            "Complete ATM Report"
        )

        props.created = datetime.datetime.now()

    def save(self):
        output = BytesIO()

        self.workbook.save(output)

        output.seek(0)

        return output

    def freeze(self):
        self.summary_sheet.freeze_panes = "A2"

        self.info_sheet.freeze_panes = "A2"

        self.month_sheet.freeze_panes = "A2"

        self.year_sheet.freeze_panes = "A2"

    from openpyxl.utils import get_column_letter

    def auto_fit(self):

        for sheet in self.workbook.worksheets:

            for column in sheet.columns:

                length = 0

                letter = get_column_letter(
                    column[0].column
                )

                for cell in column:

                    if cell.value is None:
                        continue

                    length = max(
                        length,
                        len(str(cell.value)),
                    )

                sheet.column_dimensions[
                    letter
                ].width = min(length + 4, 40)

    def build_service_cache(self):
        """
        Build service cache for ALL ATMs.

        Rules
        -----
        1. BTECH / GLOB
           - ATMServiceContract dan olinadi.
           - Eng birinchi payment mavjud bo'lgan oydan boshlab
             barcha statistik oylarga qo'llaniladi.

        2. Incassation / Rent / Electricity
           - Faqat bazadagi (year, month) yozuviga yoziladi.
           - Hech qachon keyingi oylarga tarqatilmaydi.

        3. service_month_cache
           key = (atm_id, year, month)

        4. service_year_cache
           key = (atm_id, year)
        """

        self.service_month_cache = {}
        self.service_year_cache = {}

        for atm in self.atms:

            contract = getattr(
                atm,
                "service_contract",
                None,
            )

            if not contract:
                continue

            statistics = list(
                atm.monthly_statistics.all().order_by(
                    "year",
                    "month",
                )
            )

            if not statistics:
                continue

            payments = list(
                contract.payments.all().order_by(
                    "year",
                    "month",
                    "payment_type",
                )
            )

            contract_start = None

            if payments:
                first_payment = min(
                    payments,
                    key=lambda p: (
                        p.year,
                        p.month,
                    ),
                )

                contract_start = (
                    first_payment.year,
                    first_payment.month,
                )

            # ----------------------------------------
            # Monthly cache
            # ----------------------------------------

            for stat in statistics:

                key = (
                    atm.id,
                    stat.year,
                    stat.month,
                )

                values = {
                    "btech": 0.0,
                    "glob": 0.0,
                    "incassation": 0.0,
                    "rent": 0.0,
                    "electricity": 0.0,
                }

                if (
                        contract_start
                        and (stat.year, stat.month) >= contract_start
                ):
                    values["btech"] = float(
                        contract.btech_monthly_fee or 0
                    )

                    values["glob"] = float(
                        contract.glob_monthly_fee or 0
                    )

                self.service_month_cache[key] = values

            # ----------------------------------------
            # Paymentlarni faqat o'z oyiga yozamiz
            # ----------------------------------------

            for payment in payments:

                key = (
                    atm.id,
                    payment.year,
                    payment.month,
                )

                if key not in self.service_month_cache:
                    continue

                if payment.payment_type == ATMServicePayment.PaymentType.INCASSATION:

                    self.service_month_cache[key]["incassation"] = float(
                        payment.amount
                    )

                elif payment.payment_type == ATMServicePayment.PaymentType.RENT:

                    self.service_month_cache[key]["rent"] = float(
                        payment.amount
                    )

                elif payment.payment_type == ATMServicePayment.PaymentType.ELECTRICITY:

                    self.service_month_cache[key]["electricity"] = float(
                        payment.amount
                    )

        # ----------------------------------------
        # Year cache
        # ----------------------------------------

        for (atm_id, year, month), values in self.service_month_cache.items():

            year_key = (
                atm_id,
                year,
            )

            if year_key not in self.service_year_cache:
                self.service_year_cache[year_key] = {
                    "btech": 0.0,
                    "glob": 0.0,
                    "incassation": 0.0,
                    "rent": 0.0,
                    "electricity": 0.0,
                }

            self.service_year_cache[year_key]["btech"] += values["btech"]
            self.service_year_cache[year_key]["glob"] += values["glob"]
            self.service_year_cache[year_key]["incassation"] += values["incassation"]
            self.service_year_cache[year_key]["rent"] += values["rent"]
            self.service_year_cache[year_key]["electricity"] += values["electricity"]


    def write_information(self):

        ws = self.info_sheet

        headers = [
            "Region",
            "ATM Name",
            "Address",
            "Processing",
            "ATM Model",
            "Serial Number",
            "Terminal ID",
            "Merchant ID",
            "Status",
            "Inventory Number",
            "23510 Account",
            "45265 Account",
            "Purchase Date",
            "Purchase Price",
        ]

        ws.append(headers)

        for atm in self.atms:
            tech = getattr(atm, "technical", None)

            ws.append([
                atm.region,
                atm.name,
                atm.address,
                atm.card_type,
                atm.model,
                tech.serial_number if tech else "",
                atm.terminal_id,
                tech.merchant_id if tech else "",
                tech.status if tech else "",
                tech.inventory_number if tech else "",
                tech.account_23510 if tech else "",
                tech.account_45265 if tech else "",
                tech.purchase_date if tech else "",
                float(tech.purchase_price or 0) if tech else 0,
            ])

    def write_month_statistics(self):

        ws = self.month_sheet

        ws.append([
            "Terminal ID",
            "Processing",
            "Year",
            "Month",
            "Income",
            "Cash Withdraw",
            "Repair Cost",
            "Quantity",
            "BTECH Service Monthly Fee",
            "GLOB Service Monthly Fee",
            "Incassation Payment",
            "Rent Payment",
            "Electricity Payment",
            "Status",
            "Serial Number",
            "Merchant ID",
            "Inventory Number",
            "Purchase Date",
            "Purchase Price",
            "ATM Name",
            "Region",
        ])
        totals = {
            "income": 0.0,
            "expense": 0.0,
            "repair": 0.0,
            "quantity": 0.0,
            "btech": 0.0,
            "glob": 0.0,
            "incassation": 0.0,
            "rent": 0.0,
            "electricity": 0.0,
        }


        months = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }



        atm_ids = self.atms.values_list("id", flat=True)

        statistics = (
            ATMMonthlyStatistic.objects
            .filter(atm_id__in=atm_ids)
            .select_related("atm")
            .order_by(
                "atm__region",
                "atm__terminal_id",
                "year",
                "month",
            )
        )

        for item in statistics:

            tech = self.technical_cache.get(
                item.atm.terminal_id.strip().upper()
            )

            repair = 0
            quantity = 0

            if tech:
                repairs = self.maintenance_cache.get(
                    (
                        tech.id,
                        item.year,
                        item.month,
                    ),
                    [],
                )

                repair = sum(
                    float(r.total_with_vat or 0)
                    for r in repairs
                )

                quantity = sum(
                    float(r.quantity or 0)
                    for r in repairs
                )

            service = self.service_month_cache.get(
                (
                    item.atm.id,
                    item.year,
                    item.month,
                ),
                {},
            )

            ws.append([
                item.atm.terminal_id,
                item.atm.card_type,
                item.year,
                months.get(item.month, item.month),

                float(item.income),
                float(item.expense),

                repair,
                quantity,

                float(service.get("btech", 0)),
                float(service.get("glob", 0)),

                float(service.get("incassation", 0)),
                float(service.get("rent", 0)),
                float(service.get("electricity", 0)),

                tech.status if tech else "",
                tech.serial_number if tech else "",
                tech.merchant_id if tech else "",
                tech.inventory_number if tech else "",
                tech.purchase_date if tech else "",
                float(tech.purchase_price or 0) if tech else 0,
                item.atm.name,
                item.atm.region,
            ])
            totals["income"] += float(item.income)
            totals["expense"] += float(item.expense)
            totals["repair"] += repair
            totals["quantity"] += quantity
            totals["btech"] += float(service.get("btech", 0))
            totals["glob"] += float(service.get("glob", 0))
            totals["incassation"] += float(service.get("incassation", 0))
            totals["rent"] += float(service.get("rent", 0))
            totals["electricity"] += float(service.get("electricity", 0))

        ws.append([])

        total_row = [
            "TOTAL",
            "",
            "",
            "",

            totals["income"],
            totals["expense"],

            totals["repair"],
            totals["quantity"],

            totals["btech"],
            totals["glob"],

            totals["incassation"],
            totals["rent"],
            totals["electricity"],

            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]

        ws.append(total_row)
    def write_year_statistics(self):

        ws = self.year_sheet

        ws.append([
            "Terminal ID",
            "Card Type",

            "Year",

            "Income",
            "Cash Withdraw",

            "Year Repair Cost",

            "Quantity",

            "BTECH Service Year Fee",
            "GLOB Service Year Fee",

            "Incassation Total",
            "Rent Total",
            "Electricity Total",

            "Status",
            "Serial Number",
            "Merchant ID",
            "Inventory Number",

            "Purchase Date",
            "Purchase Price",
        ])
        totals = {
            "income": 0.0,
            "expense": 0.0,
            "repair": 0.0,
            "quantity": 0.0,
            "btech": 0.0,
            "glob": 0.0,
            "incassation": 0.0,
            "rent": 0.0,
            "electricity": 0.0,
        }

        for atm in self.atms:

            statistics = (
                atm.year_statistics.all()
                .order_by("year")
            )

            tech = getattr(
                atm,
                "technical",
                None,
            )


            for item in statistics:

                repair = 0
                quantity = 0

                if tech:
                    repairs = (
                        MaintenanceItem.objects.filter(
                            technical=tech,
                            protocol_date__year=item.year,
                        )
                    )

                    repair = float(
                        repairs.aggregate(
                            total=Sum("total_with_vat")
                        )["total"] or 0
                    )

                    quantity = float(
                        repairs.aggregate(
                            total=Sum("quantity")
                        )["total"] or 0
                    )

                service = self.service_year_cache.get(
                    (
                        atm.id,
                        item.year,
                    ),
                    {},
                )
                totals["income"] += float(item.income)
                totals["expense"] += float(item.expense)
                totals["repair"] += repair
                totals["quantity"] += quantity
                totals["btech"] += float(service.get("btech", 0))
                totals["glob"] += float(service.get("glob", 0))
                totals["incassation"] += float(service.get("incassation", 0))
                totals["rent"] += float(service.get("rent", 0))
                totals["electricity"] += float(service.get("electricity", 0))

                ws.append([

                    atm.terminal_id,

                    atm.card_type,

                    item.year,

                    float(item.income),

                    float(item.expense),

                    float(repair),

                    float(quantity),

                    float(
                        service.get(
                            "btech",
                            0,
                        )
                    ),

                    float(
                        service.get(
                            "glob",
                            0,
                        )
                    ),

                    float(
                        service.get(
                            "incassation",
                            0,
                        )
                    ),

                    float(
                        service.get(
                            "rent",
                            0,
                        )
                    ),

                    float(
                        service.get(
                            "electricity",
                            0,
                        )
                    ),

                    tech.status if tech else "",

                    tech.serial_number if tech else "",

                    tech.merchant_id if tech else "",

                    tech.inventory_number if tech else "",
                    tech.purchase_date if tech else "",
                    float(tech.purchase_price or 0) if tech else 0,

                ])
            ws.append([])

        total_row = [
            "TOTAL",  # Terminal ID
            "",  # Card Type
            "",  # Year

            totals["income"],
            totals["expense"],

            totals["repair"],
            totals["quantity"],

            totals["btech"],
            totals["glob"],

            totals["incassation"],
            totals["rent"],
            totals["electricity"],

            "",  # Status
            "",  # Serial
            "",  # Merchant
            "",  # Inventory
            "",  # Purchase Date
            "",  # Purchase Price
        ]

        ws.append(total_row)
    def write_summary(self):
        atms = list(self.atms)

        atm_ids = [atm.id for atm in atms]

        technical_count = sum(
            1
            for atm in atms
            if getattr(atm, "technical", None)
        )

        active_count = sum(
            1
            for atm in atms
            if atm.is_active
        )

        uzcard_count = sum(
            1
            for atm in atms
            if atm.card_type == "UZCARD"
        )

        humo_count = sum(
            1
            for atm in atms
            if atm.card_type == "HUMO"
        )

        region_count = len(
            {
                atm.region
                for atm in atms
            }
        )

        monthly_count = ATMMonthlyStatistic.objects.filter(
            atm_id__in=atm_ids,
        ).count()

        year_count = ATMYearStatistic.objects.filter(
            atm_id__in=atm_ids,
        ).count()
        ws = self.summary_sheet

        # ======================================================
        # TITLE
        # ======================================================

        ws.merge_cells("A1:B1")
        title = ws["A1"]
        ws.row_dimensions[1].height = 30

        title.font = Font(
            size=18,
            bold=True,
            color="FFFFFF",
        )

        title.fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        title.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # ======================================================
        # REPORT INFO
        # ======================================================
        ws["A2"] = "Generated"
        ws["B2"] = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

        ws["A3"] = "Generated By"
        ws["B3"] = "Turonbank Monitoring System"

        # ======================================================
        # SUMMARY
        # ======================================================

        summary = [

            (
                "Total ATM",
                len(atms),
            ),

            (
                "Technical Linked",
                technical_count,
            ),

            (
                "Active ATM",
                active_count,
            ),

            (
                "UZCARD ATM",
                uzcard_count,
            ),

            (
                "HUMO ATM",
                humo_count,
            ),

            (
                "Regions",
                region_count,
            ),

            (
                "Monthly Statistics",
                monthly_count,
            ),

            (
                "Year Statistics",
                year_count,
            ),

        ]

        row = 8

        header_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAF7",
            end_color="D9EAF7",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for title, value in summary:
            name_cell = ws.cell(
                row=row,
                column=1,
            )

            value_cell = ws.cell(
                row=row,
                column=2,
            )

            name_cell.value = title
            value_cell.value = value

            name_cell.font = Font(bold=True)

            name_cell.fill = header_fill

            name_cell.border = border
            value_cell.border = border

            name_cell.alignment = Alignment(
                vertical="center",
            )

            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            row += 1