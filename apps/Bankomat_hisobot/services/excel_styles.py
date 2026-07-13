from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)


class ExcelStyles:
    """
    Excel uchun barcha style lar.
    """

    THIN = Side(
        style="thin",
        color="C0C0C0",
    )

    BORDER = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=THIN,
    )

    TITLE_FONT = Font(
        bold=True,
        size=18,
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
        size=11,
    )

    NORMAL_FONT = Font(
        size=10,
    )

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    SUBHEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    YEAR_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    MONTH_FILL = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
    )

    LEFT = Alignment(
        horizontal="left",
        vertical="center",
    )

    RIGHT = Alignment(
        horizontal="right",
        vertical="center",
    )

    NUMBER_FORMAT = '#,##0.000'

    