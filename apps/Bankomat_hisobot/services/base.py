from __future__ import annotations

import logging
from pathlib import Path
from abc import ABC, abstractmethod

from openpyxl import load_workbook
from django.db import transaction


logger = logging.getLogger(__name__)


class BaseExcelImporter(ABC):
    """
    Base class for all Excel importers.

    Workflow

    open workbook

        ↓

    validate

        ↓

    parse

        ↓

    bulk create/update

        ↓

    return statistics
    """

    sheet_name = None

    def __init__(self, file_path):

        self.file_path = Path(file_path)

        self.workbook = None

        self.sheet = None

        self.created = 0

        self.updated = 0

        self.skipped = 0

        self.errors = []

    ##################################################################
    # public
    ##################################################################

    def run(self):

        logger.info(
            "Import started -> %s",
            self.file_path
        )

        self.open()

        self.validate()

        with transaction.atomic():

            self.parse()

        logger.info(
            "Import finished"
        )

        return {

            "created": self.created,

            "updated": self.updated,

            "skipped": self.skipped,

            "errors": self.errors,

        }

    ##################################################################
    # workbook
    ##################################################################

    def open(self):

        if not self.file_path.exists():
            raise FileNotFoundError(self.file_path)

        self.workbook = load_workbook(

            filename=self.file_path,

            data_only=True,

            read_only=True,

        )

        if self.sheet_name:

            self.sheet = self.workbook[self.sheet_name]

        else:

            self.sheet = self.workbook.active

    ##################################################################
    # helpers
    ##################################################################

    def iter_rows(self, start=2):

        for row in self.sheet.iter_rows(

                min_row=start,

                values_only=True,

        ):

            yield row

    def clean(self, value):

        if value is None:
            return ""

        return str(value).strip()

    def is_empty(self, row):

        return all(

            cell is None or str(cell).strip() == ""

            for cell in row

        )

    ##################################################################
    # abstract
    ##################################################################

    @abstractmethod
    def validate(self):
        """
        Validate excel structure
        """

    @abstractmethod
    def parse(self):
        """
        Parse rows
        """