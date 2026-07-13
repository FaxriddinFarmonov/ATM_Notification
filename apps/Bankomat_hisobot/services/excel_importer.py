import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Any
from ..models import ATMMonthlyStatistic
from ..models import ATMYearStatistic
from decimal import Decimal
from .constants import MONTHS
from .constants import YEARS
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re
from collections import defaultdict
from .constants import (
    SHEETS,
    HEADER_ROW,
    DATA_START_ROW,
    TERMINAL_ID,
    REGION,
    ADDRESS,
    ATM_MODEL,
    ZERO,
    NOT_WORKING,
)

from ..models.ATMMonthlyStatistic import (
    ATMTURON,
    ExcelImport,
)

logger = logging.getLogger(__name__)



class ATMExcelImporter:

    def __init__(self, excel_import):

        self.excel_import = excel_import

        self.workbook = None

        self.headers = {}

        self.exist_atms = {}

        self.atm_create_list = []

        self.atm_update_list = []

        self.month_create_list = []

        self.month_update_list = []

        self.year_create_list = []

        self.year_update_list = []

        self.exist_months = {}

        self.exist_years = {}

        # NEW
        self.month_headers = []

        self.year_headers = []
        self.year_create_keys = set()





    def load_workbook(self):

        logger.info("Opening excel file...")

        self.workbook = load_workbook(
            filename=self.excel_import.file.path,
            data_only=True,
        )

        logger.info("Workbook loaded.")

    def load_existing_atms(self):

        logger.info("Loading existing ATMs...")

        self.exist_atms = {

            str(obj.terminal_id): obj

            for obj in ATMTURON.objects.all()

        }
        print(len(self.exist_atms))
        logger.info(

            "Loaded %s ATMs",

            len(self.exist_atms),

        )
        print(list(self.exist_atms.keys())[:20])



    def get_headers(
        self,
        sheet: Worksheet,
    ) -> Dict[str, int]:

        headers = {}

        for col in range(1, sheet.max_column + 1):

            value = sheet.cell(
                row=HEADER_ROW,
                column=col,
            ).value

            if value is None:
                continue

            value = str(value).strip()

            headers[value] = col

        return headers


    def get_cell(

        self,

        sheet: Worksheet,

        row: int,

        column_name: str,

    ):

        column = self.headers.get(column_name)

        if column is None:

            return None

        return sheet.cell(
            row=row,
            column=column,
        ).value

    def to_decimal(
        self,
        value,
    ) -> Decimal:

        if value is None:
            return ZERO

        if value == "":
            return ZERO

        if isinstance(value, Decimal):
            return value

        value = str(value).strip()

        if value == "":
            return ZERO

        if value == NOT_WORKING:
            return ZERO

        value = value.replace(" ", "")

        value = value.replace(",", ".")

        try:

            return Decimal(value)

        except InvalidOperation:

            return ZERO

    def clean_string(
        self,
        value,
    ) -> str:

        if value is None:

            return ""

        return str(value).strip()

    def validate_row(
            self,
            sheet,
            row,
    ) -> bool:
        """
        Satr import qilinishga yaroqliligini tekshiradi.
        """

        try:
            terminal = self.get_cell(
                sheet=sheet,
                row=row,
                column_name=TERMINAL_ID,
            )
        except Exception:
            return False

        if terminal is None:
            return False

        terminal = self.clean_string(terminal)

        if not terminal:
            return False

        # Excelda ba'zan bunday qiymatlar uchraydi
        if terminal in (
                "-",
                "--",
                "None",
                "NULL",
                "nan",
        ):
            return False

        # Terminal ID uzunligi juda kichik bo'lsa ham o'tkazmaymiz
        if len(terminal) < 3:
            return False

        return True




    def parse_atm(
            self,
            sheet,
            row,
            card_type,
    ):
        """
        Exceldagi bitta ATM satrini parse qiladi.
        """

        terminal = self.clean_string(
            self.get_cell(
                sheet,
                row,
                TERMINAL_ID,
            )
        )

        # Terminal ID bo'lmasa satrni tashlab ketamiz
        if not terminal:
            return

        region = self.clean_string(
            self.get_cell(
                sheet,
                row,
                REGION,
            )
        )

        address = self.clean_string(
            self.get_cell(
                sheet,
                row,
                ADDRESS,
            )
        )

        model = self.clean_string(
            self.get_cell(
                sheet,
                row,
                ATM_MODEL,
            )
        )

        name = address

        card_type = card_type.upper()

        if terminal in self.exist_atms:

            atm = self.exist_atms[terminal]

            changed = False

            if atm.region != region:
                atm.region = region
                changed = True

            if atm.address != address:
                atm.address = address
                changed = True

            if atm.name != name:
                atm.name = name
                changed = True

            if atm.model != model:
                atm.model = model
                changed = True

            if atm.card_type != card_type:
                atm.card_type = card_type
                changed = True

            if changed:
                self.atm_update_list.append(atm)

        else:

            atm = ATMTURON(
                terminal_id=terminal,
                region=region,
                name=name,
                address=address,
                model=model,
                card_type=card_type,
            )

            self.atm_create_list.append(atm)
            self.exist_atms[terminal] = atm
        return atm



    @transaction.atomic
    def run(self):

        self.load_workbook()
        self.load_existing_atms()
        for sheet in SHEETS:
            self.parse_sheet_only_atms(sheet)

        self.save_atms()

        self.load_existing_atms()

        self.load_existing_months()

        self.load_existing_years()  # <-- SHU YERGA KO'CHDI

        for sheet in SHEETS:
            self.parse_sheet_statistics(sheet)

        self.save_month_statistics()
        self.rebuild_year_statistics()

        self.save_year_statistics()



    def save_atms(self):

        ...

        if self.atm_create_list:
            ATMTURON.objects.bulk_create(
                self.atm_create_list,
                batch_size=500,
            )

        if self.atm_update_list:
            ATMTURON.objects.bulk_update(
                self.atm_update_list,
                [
                    "region",
                    "address",
                    "model",
                    "card_type",
                ],
                batch_size=500,
            )

        self.atm_create_list.clear()
        self.atm_update_list.clear()

    def clean_decimal(self, value):

        if value is None:
            return Decimal("0")

        value = str(value)

        value = value.replace(" ", "")

        value = value.replace(",", ".")

        if value == "":
            return Decimal("0")

        try:

            return Decimal(value)

        except Exception:

            return Decimal("0")

    def parse_month_statistics(
            self,
            sheet,
            row,
            atm,
    ):
        """
        Exceldagi barcha oylik statistikalarni avtomatik parse qiladi.

        MONTHS konstantasiga bog'liq emas.
        Excelda qaysi yil va oy mavjud bo'lsa,
        hammasini bazaga saqlaydi.
        """

        for item in self.month_headers:

            year = item["year"]
            month = item["month"]

            expense = self.clean_decimal(
                self.get_cell(
                    sheet,
                    row,
                    item["expense"],
                )
            )

            income = self.clean_decimal(
                self.get_cell(
                    sheet,
                    row,
                    item["income"],
                )
            )

            key = (
                atm.terminal_id,
                year,
                month,
            )

            if key in self.exist_months:

                stat = self.exist_months[key]

                changed = False

                if stat.income != income:
                    stat.income = income
                    changed = True

                if stat.expense != expense:
                    stat.expense = expense
                    changed = True

                if changed:
                    self.month_update_list.append(stat)

                continue

            stat = ATMMonthlyStatistic(
                atm=atm,
                year=year,
                month=month,
                income=income,
                expense=expense,
            )

            self.month_create_list.append(stat)

            self.exist_months[key] = stat
    def load_existing_months(self):

        logger.info("Loading monthly statistics...")

        self.exist_months = {}

        for obj in ATMMonthlyStatistic.objects.select_related("atm"):
            key = (
                obj.atm.terminal_id,
                obj.year,
                obj.month,
            )

            self.exist_months[key] = obj

        logger.info(
            "Loaded %s monthly statistics",
            len(self.exist_months),
        )

    def save_month_statistics(self):
        """
        Oylik statistikalarni bazaga saqlaydi.
        """


        if self.month_create_list:
            ATMMonthlyStatistic.objects.bulk_create(
                self.month_create_list,
                batch_size=1000,
            )

        if self.month_update_list:
            ATMMonthlyStatistic.objects.bulk_update(
                self.month_update_list,
                [
                    "expense",
                    "income",
                ],
                batch_size=1000,
            )

    def rebuild_year_statistics(self):
        """
        Oylik statistikalar asosida yillik statistikalarni qayta hisoblaydi.

        Exceldagi "Jami" ustunlariga bog'liq emas.
        """

        logger.info("=" * 80)
        logger.info("REBUILD YEAR STATISTICS FROM MONTHS")
        logger.info("=" * 80)

        grouped = defaultdict(
            lambda: {
                "income": 0,
                "expense": 0,
            }
        )

        #
        # Avval mavjud bazadagi oyliklar
        #
        for stat in self.exist_months.values():
            key = (
                stat.atm.terminal_id,
                stat.year,
            )

            grouped[key]["income"] += stat.income or 0
            grouped[key]["expense"] += stat.expense or 0

        #
        # Keyin yangi create bo'ladigan oylar
        #
        for stat in self.month_create_list:
            key = (
                stat.atm.terminal_id,
                stat.year,
            )

            grouped[key]["income"] += stat.income or 0
            grouped[key]["expense"] += stat.expense or 0

        #
        # Har bir yilni update/create qilamiz
        #
        for (terminal, year), total in grouped.items():

            atm = self.exist_atms[terminal]

            key = (
                terminal,
                year,
                atm.card_type,
            )

            if key in self.exist_years:

                obj = self.exist_years[key]

                changed = False

                if obj.income != total["income"]:
                    obj.income = total["income"]
                    changed = True

                if obj.expense != total["expense"]:
                    obj.expense = total["expense"]
                    changed = True

                if changed:
                    self.year_update_list.append(obj)

            else:

                obj = ATMYearStatistic(
                    atm=atm,
                    year=year,
                    card_type=atm.card_type,
                    income=total["income"],
                    expense=total["expense"],
                )

                self.year_create_list.append(obj)

                self.exist_years[key] = obj

        logger.info(
            "Year statistics rebuilt: CREATE=%s UPDATE=%s",
            len(self.year_create_list),
            len(self.year_update_list),
        )

    def parse_sheet_only_atms(self, sheet_name):
        """
        Faqat ATMlarni parse qiladi.
        """

        logger.info("=" * 80)
        logger.info("Reading ATM sheet: %s", sheet_name)
        logger.info("=" * 80)

        if sheet_name not in self.workbook.sheetnames:
            return

        sheet = self.workbook[sheet_name]

        self.headers = self.get_headers(sheet)

        for row in range(DATA_START_ROW, sheet.max_row + 1):

            if not self.validate_row(sheet, row):
                continue

            self.parse_atm(
                sheet=sheet,
                row=row,
                card_type=sheet_name,
            )

    def parse_sheet_statistics(self, sheet_name):
        """
        Faqat oylik/yillik statistikalarni parse qiladi.
        """

        logger.info("=" * 80)
        logger.info("Reading statistics sheet: %s", sheet_name)
        logger.info("=" * 80)

        if sheet_name not in self.workbook.sheetnames:
            return

        sheet = self.workbook[sheet_name]

        self.headers = self.get_headers(sheet)
        self.build_statistics_headers()
        for row in range(DATA_START_ROW, sheet.max_row + 1):

            if not self.validate_row(sheet, row):
                continue

            terminal = self.clean_string(
                self.get_cell(
                    sheet,
                    row,
                    TERMINAL_ID,
                )
            )

            atm = self.exist_atms.get(terminal)

            if atm is None:
                continue

            self.parse_month_statistics(
                sheet=sheet,
                row=row,
                atm=atm,

            )

            self.parse_year_statistics(
                sheet=sheet,
                row=row,
                atm=atm,
                card_type=sheet_name.upper(),

            )
        print("=" * 80)
        print("MONTH HEADERS")

        for item in self.month_headers:
            print(item)

        print("=" * 80)
        print("YEAR HEADERS")

        for item in self.year_headers:
            print(item)

    def load_existing_years(self):
        """
        Bazadagi barcha yillik statistikalarni RAM ga yuklaydi.
        """

        logger.info("=" * 80)
        logger.info("LOAD EXISTING YEAR STATISTICS")
        logger.info("=" * 80)

        self.exist_years = {}

        for obj in ATMYearStatistic.objects.select_related("atm"):
            key = (
                obj.atm.terminal_id,
                obj.year,
                obj.card_type,
            )

            self.exist_years[key] = obj

        logger.info(
            "Loaded %s yearly statistics",
            len(self.exist_years),
        )

    def parse_year_statistics(
            self,
            sheet,
            row,
            atm,
            card_type,
    ):
        """
        Exceldagi yillik statistikalarni parse qiladi.

        Endi bu metod YEARS konstantasiga bog'liq emas.
        Excelda nechta yil bo'lsa, hammasini avtomatik o'qiydi.
        """

        for item in self.year_headers:

            year = item["year"]

            expense = self.clean_decimal(
                self.get_cell(
                    sheet,
                    row,
                    item["expense"],
                )
            )

            income = self.clean_decimal(
                self.get_cell(
                    sheet,
                    row,
                    item["income"],
                )
            )
            if (income or 0) == 0 and (expense or 0) == 0:
                income, expense = self.calculate_year_total(
                    atm,
                    year,
                )

            print(
                "YEAR:",
                atm.terminal_id,
                year,
                income,
                expense,
            )

            key = (
                atm.terminal_id,
                year,
                card_type,
            )

            if key in self.exist_years:

                stat = self.exist_years[key]

                changed = False

                if stat.expense != expense:
                    stat.expense = expense
                    changed = True

                if stat.income != income:
                    stat.income = income
                    changed = True

                if changed:
                    self.year_update_list.append(stat)

                continue

            duplicate_key = (
                atm.terminal_id,
                year,
                card_type,
            )

            if key  in self.year_create_keys:
                continue

            stat = ATMYearStatistic(
                atm=atm,
                year=year,
                card_type=card_type,
                expense=expense,
                income=income,
            )

            self.year_create_list.append(stat)
            self.year_create_keys.add(key)

            self.exist_years[key] = stat

    def save_year_statistics(self):
        """
        Yillik statistikalarni bazaga saqlaydi.
        """

        logger.info("=" * 80)
        logger.info("SAVE YEAR STATISTICS")
        logger.info("CREATE: %s", len(self.year_create_list))
        logger.info("UPDATE: %s", len(self.year_update_list))
        logger.info("=" * 80)

        if self.year_create_list:
            ATMYearStatistic.objects.bulk_create(
                self.year_create_list,
                batch_size=500,
            )

        if self.year_update_list:

            unique = {}

            for obj in self.year_update_list:
                unique[obj.pk] = obj

            ATMYearStatistic.objects.bulk_update(
                list(unique.values()),
                [
                    "income",
                    "expense",
                ],
                batch_size=500,
            )

    def build_statistics_headers(self):
        """
        Excel headerlarini avtomatik tahlil qiladi.

        - barcha yillarni topadi
        - barcha oylarni topadi
        - MONTHS va YEARS konstantalariga bog'liq emas
        """

        logger.info("=" * 80)
        logger.info("BUILD STATISTICS HEADERS")
        logger.info("=" * 80)

        self.month_headers = []
        self.year_headers = []

        month_names = {
            "январ": 1,
            "феврал": 2,
            "март": 3,
            "апрел": 4,
            "апрель": 4,
            "май": 5,
            "июн": 6,
            "июнь": 6,
            "июл": 7,
            "июль": 7,
            "август": 8,
            "сентябр": 9,
            "сентабр": 9,
            "октябр": 10,
            "октабр": 10,
            "ноябр": 11,
            "декабр": 12,
        }

        month_map = {}
        year_map = {}

        for header in self.headers.keys():

            if not header:
                continue

            text = str(header).strip().lower()

            year_match = re.search(r"20\d{2}", text)

            if not year_match:
                continue

            year = int(year_match.group())

            # ---------------- YEAR -----------------

            if "жами" in text:

                item = year_map.setdefault(
                    year,
                    {
                        "year": year,
                        "income": None,
                        "expense": None,
                    },
                )

                if "даромад" in text:
                    item["income"] = header

                if "чиқим" in text:
                    item["expense"] = header

                continue

            # ---------------- MONTH -----------------

            month = None

            for month_name, month_number in month_names.items():

                if month_name in text:
                    month = month_number
                    break

            if month is None:
                continue

            item = month_map.setdefault(
                (year, month),
                {
                    "year": year,
                    "month": month,
                    "income": None,
                    "expense": None,
                },
            )

            if "даромад" in text:
                item["income"] = header

            if "чиқим" in text:
                item["expense"] = header

        self.month_headers = sorted(
            month_map.values(),
            key=lambda x: (x["year"], x["month"]),
        )

        self.year_headers = sorted(
            year_map.values(),
            key=lambda x: x["year"],
        )

        logger.info("=" * 80)
        logger.info("MONTH HEADERS (%s)", len(self.month_headers))

        for item in self.month_headers:
            logger.info(item)

        logger.info("=" * 80)
        logger.info("YEAR HEADERS (%s)", len(self.year_headers))

        for item in self.year_headers:
            logger.info(item)

        print("=" * 80)
        print("ALL HEADERS")

        for header in self.headers.keys():
            print(header)

    def calculate_year_total(self, atm, year):

        income = 0
        expense = 0

        for stat in self.exist_months.values():

            if (
                    stat.atm.terminal_id == atm.terminal_id
                    and stat.year == year
            ):
                income += stat.income or 0
                expense += stat.expense or 0

        print(
            "TOTAL:",
            atm.terminal_id,
            year,
            income,
            expense,
        )

        return income, expense