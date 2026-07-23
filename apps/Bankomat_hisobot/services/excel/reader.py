from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:

    SHEET_NAME = "SERVIZ tulovlar"

    def __init__(self, file_path):

        self.file_path = Path(file_path)

        self.workbook = None

        self.sheet = None

    def open(self):

        self.workbook = load_workbook(
            filename=self.file_path,
            data_only=True,
        )

        target = self.SHEET_NAME.strip().lower()

        for sheet in self.workbook.worksheets:

            if sheet.title.strip().lower() == target:

                self.sheet = sheet

                return sheet

        raise ValueError(
            f"'{self.SHEET_NAME}' sheet topilmadi.\n"
            f"Mavjud sheetlar: "
            f"{self.workbook.sheetnames}"
        )