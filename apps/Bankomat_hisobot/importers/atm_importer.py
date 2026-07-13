from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from ..models.ATMMonthlyStatistic import ATMTURON
from ..models.full_models import ATMTechnical


class ATMExcelImporter:

    SHEET_NAME = "FULL"

    HEADER_ROW = 1
    START_ROW = 2

    def __init__(self, file):
        self.file = Path(file)

        self.workbook = None
        self.sheet = None

        self.headers = {}

        self.objects_create = []
        self.objects_update = []

        self.exists = {}

    def run(self):

        self.load_workbook()

        self.get_sheet()

        self.get_headers()

        self.prepare_objects()

        result = self.bulk_save()

        return result

    def load_workbook(self):
        self.workbook = load_workbook(
            self.file,
            data_only=True
        )

    def get_sheet(self):
        self.sheet = self.workbook[self.SHEET_NAME]

    def get_headers(self):
        self.headers = {}

        for cell in self.sheet[self.HEADER_ROW]:
            if cell.value:
                self.headers[str(cell.value).strip()] = cell.column


    def value(self, row, name):

        column = self.headers[name]

        value = self.sheet.cell(
            row=row,
            column=column
        ).value

        if value is None:
            return ""

        return str(value).strip()

    def parse_row(self, row):

        terminal_id = self.value(row, "Terminal ID")

        if not terminal_id:
            return None

        obj = ATMTechnical(
            card_type=self.value(row, "Bankomat turi").strip().upper(),
            model_name=self.value(row, "Bankomat modeli"),
            status=self.value(row, "Holati").strip().upper(),
            serial_number=self.value(row, "Seriya raqami"),
            inventory_number=self.value(row, "Inventar raqami"),
            address=self.value(row, "Yuridik manzili"),
            merchant_id=self.value(row, "Merchant ID"),
            terminal_id=terminal_id,
            account_23510=self.value(row, "23510"),
            account_45265=self.value(row, "45265"),
        )

        obj.atm = self.turons.get(terminal_id)

        return obj

    def prepare_objects(self):

        existing = ATMTechnical.objects.in_bulk(field_name="terminal_id")
        self.turons = {
            atm.terminal_id: atm
            for atm in ATMTURON.objects.exclude(terminal_id="")
        }
        seen = set()

        for row in range(self.START_ROW, self.sheet.max_row + 1):

            obj = self.parse_row(row)

            if obj is None:
                continue

            if obj.terminal_id in seen:
                continue

            seen.add(obj.terminal_id)

            old = existing.get(obj.terminal_id)

            if old is None:
                self.objects_create.append(obj)
            else:
                old.card_type = obj.card_type
                old.model_name = obj.model_name
                old.status = obj.status
                old.serial_number = obj.serial_number
                old.inventory_number = obj.inventory_number
                old.address = obj.address
                old.merchant_id = obj.merchant_id
                old.account_23510 = obj.account_23510
                old.account_45265 = obj.account_45265

                old.atm = obj.atm

                self.objects_update.append(old)

    from django.db import transaction

    @transaction.atomic
    def bulk_save(self):

        created = len(self.objects_create)
        updated = len(self.objects_update)

        if self.objects_create:
            ATMTechnical.objects.bulk_create(
                self.objects_create,
                batch_size=1000
            )

        if self.objects_update:
            ATMTechnical.objects.bulk_update(
                self.objects_update,
                fields=[
                    "atm",
                    "card_type",
                    "model_name",
                    "status",
                    "serial_number",
                    "inventory_number",
                    "address",
                    "merchant_id",
                    "account_23510",
                    "account_45265",
                ],
                batch_size=1000
            )

        return {
            "created": created,
            "updated": updated,
        }

